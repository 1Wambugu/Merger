import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment
from copy import copy

# Load the workbook
# input_workbook_path = 'management.xlsx'
wb = openpyxl.load_workbook('management.xlsx')

# Create a new workbook for the merged data
merged_wb = openpyxl.Workbook()
merged_ws = merged_wb.active
merged_ws.title = 'MergedData'

# Function to copy cell styles
def copy_cell_styles(source_cell, target_cell):
    target_cell.fill = copy(source_cell.fill)
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)

# Variable to track the current row in the merged worksheet
current_row = 1

# Loop through each sheet in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]

    # Copy all rows from the current sheet to the merged sheet
    for row in sheet.iter_rows():
        for col_idx, cell in enumerate(row, start=1):
            new_cell = merged_ws.cell(row=current_row, column=col_idx, value=cell.value)
            copy_cell_styles(cell, new_cell)
        current_row += 1

# Save the merged workbook
output_workbook_path = 'merged_management.xlsx'
merged_wb.save(output_workbook_path)

print(f"Merged workbook saved at {output_workbook_path}")


